"""Ingest REAL DOL H-1B LCA disclosure data into the project's schema.

Input : a real LCA disclosure .xlsx in data/raw/ (download via browser from
        https://www.dol.gov/agencies/eta/foreign-labor/performance -- DOL's CDN
        blocks scripted downloads, so this one file is fetched by hand).
Output: data/market_comp.csv  (role_family, level, metro, years_exp, base_salary)
        -- the exact schema src/model.py trains on, so nothing downstream changes.

What it does, row by row (streamed via openpyxl read_only so the ~180MB annual
file never has to fit in memory as a DataFrame):
  1. keep only CERTIFIED applications,
  2. keep only tech SOC codes we can map to a role family,
  3. normalize WAGE_RATE_OF_PAY_FROM to an ANNUAL base using WAGE_UNIT_OF_PAY,
  4. derive role_family from SOC (refined by JOB_TITLE keywords),
  5. derive level from the free-text JOB_TITLE via src/leveling,
  6. bucket the worksite city/state into one of our METROS,
  7. drop implausible wages.

IMPORTANT CAVEAT (documented in the README too): the LCA wage is the *offered
base wage* on the petition -- it excludes equity and bonus and clusters near
prevailing-wage floors. So these real bands read LOWER than total-comp sources
like Levels.fyi. That's a data-provenance feature to call out, not a bug.
"""

import os
import sys
import glob
import csv

from openpyxl import load_workbook

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.leveling import normalize_title  # noqa: E402
from src.config import LEVEL_YEARS  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "raw")

# --- SOC code -> our role family (2018 SOC). Prefixes/exact codes we trust. ---
SOC_TO_ROLE = {
    "15-1252": "Software Engineer",   # Software Developers
    "15-1251": "Software Engineer",   # Computer Programmers
    "15-1254": "Software Engineer",   # Web Developers
    "15-1255": "Software Engineer",   # Web & Digital Interface Designers (often SWE)
    "15-1211": "Software Engineer",   # Computer Systems Analysts
    "15-2051": "Data Scientist",      # Data Scientists
    "15-1242": "Data Engineer",       # Database Administrators
    "15-1243": "Data Engineer",       # Database Architects
    "15-1245": "Data Engineer",       # Database/Network Admins & Architects
    "11-3021": "Engineering Manager", # Computer & Information Systems Managers
}

# Title keywords that should *override* the SOC-derived role family.
_TITLE_ROLE_OVERRIDE = [
    ("Engineering Manager", ["engineering manager", "eng manager", "manager, software"]),
    ("Data Scientist", ["data scientist", "machine learning", " ml ", "applied scientist", "research scientist"]),
    ("Data Engineer", ["data engineer", "analytics engineer", "etl"]),
    ("Product Manager", ["product manager", "product owner"]),
    ("Product Designer", ["ux designer", "ui designer", "product designer", "user experience"]),
]

WAGE_TO_ANNUAL = {
    "year": 1.0, "yr": 1.0, "annual": 1.0,
    "hour": 2080.0, "hr": 2080.0,
    "week": 52.0, "wk": 52.0,
    "bi-weekly": 26.0, "biweekly": 26.0,
    "month": 12.0, "mth": 12.0, "mo": 12.0,
}

# City -> metro bucket (lowercased). Anything unmatched falls back by state.
_CITY_METRO = {}
for metro, cities in {
    "SF Bay Area": ["san francisco", "san jose", "mountain view", "palo alto", "sunnyvale",
                    "santa clara", "menlo park", "cupertino", "redwood city", "south san francisco",
                    "fremont", "oakland", "berkeley", "foster city", "san mateo"],
    "New York City": ["new york", "brooklyn", "manhattan", "new york city", "long island city"],
    "Seattle": ["seattle", "bellevue", "redmond", "kirkland"],
    "Boston": ["boston", "cambridge"],
    "Austin": ["austin"],
    "Denver": ["denver", "boulder"],
}.items():
    for c in cities:
        _CITY_METRO[c] = metro

# State -> coarse metro fallback for cities we don't recognize.
_STATE_METRO = {"CA": "Other US Metro", "NY": "Other US Metro", "WA": "Other US Metro"}

CERTIFIED = {"certified", "certified - withdrawn"}
MIN_ANNUAL, MAX_ANNUAL = 40_000, 1_500_000


def _to_role(soc_code, title):
    t = f" {str(title).lower()} "
    for role, kws in _TITLE_ROLE_OVERRIDE:
        if any(kw in t for kw in kws):
            return role
    soc = str(soc_code).strip()[:7]
    return SOC_TO_ROLE.get(soc)


def _to_metro(city, state):
    c = str(city).strip().lower()
    if c in _CITY_METRO:
        return _CITY_METRO[c]
    return _STATE_METRO.get(str(state).strip().upper(), "Other US Metro")


def _to_annual(wage_from, unit):
    try:
        w = float(str(wage_from).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None
    mult = WAGE_TO_ANNUAL.get(str(unit).strip().lower())
    if mult is None or w <= 0:
        return None
    return w * mult


def _find_file():
    cands = [f for f in glob.glob(os.path.join(RAW, "*.xlsx"))
             if "lca" in os.path.basename(f).lower()]
    if not cands:
        cands = glob.glob(os.path.join(RAW, "*.xlsx"))
    return cands[0] if cands else None


def ingest(path=None, out=None):
    path = path or _find_file()
    if not path or not os.path.exists(path):
        sys.exit(
            "No LCA .xlsx found in data/raw/.\n"
            "Download it (browser) from https://www.dol.gov/agencies/eta/foreign-labor/performance\n"
            "  -> 'LCA Programs (H-1B, H-1B1, E-3)' -> latest FY disclosure file\n"
            "and save it into data/raw/, then re-run this script."
        )
    out = out or os.path.join(HERE, "market_comp.csv")
    print(f"Reading {os.path.basename(path)} (streaming)...")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().upper() if h is not None else "" for h in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    def col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    c_status = col("CASE_STATUS", "STATUS")
    c_title = col("JOB_TITLE")
    c_soc = col("SOC_CODE", "SOC_CD")
    c_wage = col("WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_FROM_1")
    c_unit = col("WAGE_UNIT_OF_PAY", "WAGE_UNIT_OF_PAY_1")
    c_city = col("WORKSITE_CITY", "WORKSITE_CITY_1")
    c_state = col("WORKSITE_STATE", "WORKSITE_STATE_1")
    missing = [n for n, v in {"CASE_STATUS": c_status, "JOB_TITLE": c_title,
               "SOC_CODE": c_soc, "WAGE_RATE_OF_PAY_FROM": c_wage,
               "WAGE_UNIT_OF_PAY": c_unit}.items() if v is None]
    if missing:
        sys.exit(f"Expected columns not found: {missing}\nHeader was: {header[:25]}")

    kept = 0
    seen = 0
    stats = {}
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["role_family", "level", "metro", "years_exp", "base_salary"])
        for r in rows:
            seen += 1
            if c_status is not None:
                if str(r[c_status]).strip().lower() not in CERTIFIED:
                    continue
            role = _to_role(r[c_soc], r[c_title])
            if role is None:
                continue
            annual = _to_annual(r[c_wage], r[c_unit])
            if annual is None or not (MIN_ANNUAL <= annual <= MAX_ANNUAL):
                continue
            _, level = normalize_title(r[c_title])
            metro = _to_metro(r[c_city] if c_city is not None else "",
                              r[c_state] if c_state is not None else "")
            years = LEVEL_YEARS[level]  # H-1B has no tenure field -> impute from level
            w.writerow([role, level, metro, years, round(annual, -2)])
            kept += 1
            stats[role] = stats.get(role, 0) + 1
            if seen % 100_000 == 0:
                print(f"  scanned {seen:,} rows, kept {kept:,}...")
    wb.close()

    with open(os.path.join(HERE, "market_source.txt"), "w") as f:
        f.write(f"REAL: DOL H-1B LCA disclosure ({os.path.basename(path)}), "
                f"{kept:,} certified records; offered BASE wage only (excl. equity/bonus)")

    print(f"\nDone. Scanned {seen:,} rows, kept {kept:,} -> {os.path.relpath(out)}")
    print("By role family:")
    for role, n in sorted(stats.items(), key=lambda kv: -kv[1]):
        print(f"  {role:20s} {n:>8,}")
    if kept < 200:
        print("\nWARNING: very few rows kept -- check the column layout / SOC filter.")


if __name__ == "__main__":
    ingest()
